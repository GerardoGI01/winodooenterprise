# -*- coding: utf-8 -*-

import io
import base64
import xlwt
import openpyxl
from babel.numbers import format_currency
from openpyxl.utils import get_column_letter

from odoo import fields, models, api, _
from odoo.exceptions import UserError, ValidationError
from odoo.tools import float_compare, float_is_zero


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    def generate_excel_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        # Recolectar reglas salariales únicas
        unique_salary_rules = {}
        for payslip in self.slip_ids:
            for line in payslip.line_ids:
                if line.salary_rule_id:
                    rule_name = line.salary_rule_id.name
                    rule_sequence = line.salary_rule_id.sequence
                    if rule_name not in unique_salary_rules:
                        unique_salary_rules[rule_name] = {
                            'sequence': rule_sequence,
                            'line_id': line.salary_rule_id,
                        }

        sorted_rules = sorted(
            unique_salary_rules.values(),
            key=lambda x: (x['sequence'], x['line_id'].name)
        )

        # Encabezados
        headers = [
            'Empleado',
            'Número de Cédula',
            'Fecha de Ingreso',
            'Fecha Finalización',
            'Centro de Costos',
        ] + [rule['line_id'].name for rule in sorted_rules]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
            ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)

            max_length = 0
            column = get_column_letter(col_num)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        ws.freeze_panes = 'B1'

        # Llenar filas
        row_num = 2
        for payslip in self.slip_ids:
            employee_name = payslip.employee_id.name
            employee_id = payslip.employee_id.identification_id
            date_of_joining = payslip.employee_id.date_of_joining
            date_of_end = payslip.employee_id.contract_id.date_end
            account_cost = payslip.employee_id.contract_id.analytic_account_id.name
            date_end = date_of_end if date_of_end else 0

            ws.cell(row=row_num, column=1, value=employee_name)
            ws.cell(row=row_num, column=2, value=employee_id)
            ws.cell(row=row_num, column=3, value=date_of_joining)
            ws.cell(row=row_num, column=4, value=date_end)
            ws.cell(row=row_num, column=5, value=account_cost)

            ws.cell(row=row_num, column=1).font = openpyxl.styles.Font(bold=True)

            totals_by_rule = {rule['line_id'].name: 0.0 for rule in sorted_rules}

            for line in payslip.line_ids:
                if line.salary_rule_id and line.salary_rule_id.name in totals_by_rule:
                    totals_by_rule[line.salary_rule_id.name] += line.amount

            for col_num, rule_data in enumerate(sorted_rules, 6):
                rule_name = rule_data['line_id'].name
                total_amount = totals_by_rule[rule_name]
                ws.cell(row=row_num, column=col_num, value=total_amount)

            row_num += 1

        # Guardar archivo en binario
        archivo = io.BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        data = archivo.read()

        if data:
            file_id = self.env['file.imp'].create(
                {'filecontent': base64.b64encode(data)}
            )
            filename_field = 'Informe_Recibo_Nomina'
            if file_id and file_id.id:
                return {
                    'res_model': 'ir.actions.act_url',
                    'type': 'ir.actions.act_url',
                    'target': 'new',
                    'url': (
                        'web/content/?model=file.imp&id={0}'
                        '&filename_field={1}'
                        '&field=filecontent&download=true'
                        '&filename={1}.xlsx'.format(
                            file_id.id,
                            filename_field,
                        )
                    ),
                }
        else:
            raise ValueError('Error de Descarga - No se pudo generar el archivo solicitado.')
